import pandas as pd
import openpyxl
import os


def collect_data(path):
    headers = {'Account List': ['Management Account ID', 'Management Name', 'Management Company ALN ID',
                                'Management Owner Name', 'Sales Region', 'Manager Name', 'Validated?'],
               'Property List': ['Management Account ID', 'Management Name', 'Management Company ALN ID',
                                 'Management Owner Name', 'Sales Region', 'Manager Name', 'Property Account ID',
                                 'Property Name', 'Property ALN ID', 'Current Property Units', 'Current Property Type',
                                 'New Property Units', 'New Property Type'],
               'COM Properties': ['Property Account ID', 'Property Name', 'New Management Company ID',
                                  'New Management Company Name'],
               'Property Add': ['Property Name', 'Property Type', 'Property Units', 'Shipping Street Address',
                                'Shipping City', 'Shipping State', 'Shipping Zip', 'Billing Street Address', 'Billing City',
                                'Billing State', 'Billing Zip', 'Management Account ID', 'Management Account Name',
                                'Property ALN ID'],
               'Duplicate Properties': ['Property Account ID In Book', 'Property Name In Book',
                                        'Duplicate Property Account ID', 'Duplicate Property Name']
               }

    # create a dictionary of blank lists for each sheet name from the headers file
    data_values = {}
    for key in headers.keys():
        data_values[key] = []

    # loop through each workbook in the path, if the sheet isn't in the header variance, skip that sheet
    for file in os.listdir(path):
        wrkbk = openpyxl.load_workbook(f'{path}/{file}')
        for sheet in wrkbk.worksheets:
            if sheet.title not in headers:
                continue
            num_of_rows = sheet.max_row

            # iterate over each row in the given sheet based on the max number of rows and the number of columns
            # in the header dictionary for that sheet
            for rows in sheet.iter_rows(min_row=2, max_row=num_of_rows, min_col=1, max_col=len(headers[sheet.title]),
                                        values_only=True):
                data_values[sheet.title].append(list(rows))

    # for each collected key and value in the complete dictionary, create a dataframe by rows
    # drop any completely blank rows and export each dataframe to a csv file
    for k, v in headers.items():
        df = pd.DataFrame(data_values[k], columns=v)
        df.dropna(how='all', inplace=True)
        df.to_csv(f'{k}.csv')
