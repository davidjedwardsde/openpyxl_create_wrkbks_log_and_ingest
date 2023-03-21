import openpyxl
from openpyxl.styles.protection import Protection
from openpyxl.utils.dataframe import dataframe_to_rows
from openpyxl.utils import get_column_letter
from openpyxl.worksheet.datavalidation import DataValidation
from openpyxl.styles import PatternFill
import pandas as pd
import os


def create_wrkbks(path):
    ext = ('.xlsx')
    for file in os.listdir(path):
        if file.endswith(ext):
            print(f'Processing {file}...')
    
            # headers for all of the sheets
            com_prop_headers = ['Property Account ID', 'Property Name', 'New Management Company ID',
                                'New Management Company Name']
            prop_add_headers = ['Property Name', 'Property Type', 'Property Units',
                                'Shipping Street Address', 'Shipping City', 'Shipping State', 'Shipping Zip',
                                'Billing Street Address', 'Billing City', 'Billing State', 'Billing Zip',
                                'Management Account ID', 'Management Account Name', 'Property ALN ID']
            dup_prop_headers = ['Property Account ID In Book', 'Property Name In Book',
                                'Duplicate Property Account ID', 'Duplicate Property Name']
    
            # load workbook - will need to make this dynamic in function
            wb = openpyxl.load_workbook(os.path.join(path, file))
    
            # rename existing sheet to Property List
            ws2 = wb.active
            ws2.title = 'Property List'
    
            # create new sheets
            ws1 = wb.create_sheet('Account List', 0)
            ws3 = wb.create_sheet('COM Properties', 2)
            ws4 = wb.create_sheet('Property Add', 3)
            ws5 = wb.create_sheet('Duplicate Properties', 4)
    
            # Add headers to respective sheets
            ws3.append(com_prop_headers)
            ws4.append(prop_add_headers)
            ws5.append(dup_prop_headers)
    
            # load property list to pandas df, drop dupes, and convert to sheet
            prop_df = pd.DataFrame(ws2.values)
            prop_df.drop(prop_df.columns[[6, 7, 8, 9, 10, 11, 12]], axis=1,
                         inplace=True)  # convert back to names - be explicit, magic numbers issue
            prop_df.drop_duplicates(subset=[0], inplace=True)
    
            for r in dataframe_to_rows(prop_df, index=False, header=False):
                ws1.append(r)
    
            # add Validated column, and data validation
            ws1['G1'] = 'Validated?'
            ws1_max_row = ws1.max_row
            ws1_data_val = DataValidation(type='list', formula1='"Yes,No"', allowBlank=True, showDropDown=False)
            ws1.add_data_validation(ws1_data_val)
            dv_app = "G2:G" + str(ws1.max_row)
            ws1_data_val.add(dv_app)
    
            # add validation for property types for property types
            ws2_max_row = ws2.max_row
            ws2_data_val = DataValidation(type='list',
                                          formula1='"Affordable, Conventional, Military, NYC, Senior, Student"',
                                          allowBlank=True, showDropDown=False)
            ws2.add_data_validation(ws2_data_val)
            dv_app2 = 'M2:M' + str(ws2_max_row)
            ws2_data_val.add(dv_app2)
    
            # autofit cols on worksheet 1
            for idx, col in enumerate(ws1.columns, 1):
                ws1.column_dimensions[get_column_letter(idx)].auto_size = True
    
            # autofit cols on worksheet 2
            for idx, col in enumerate(ws2.columns, 1):
                ws2.column_dimensions[get_column_letter(idx)].auto_size = True
    
            # fill background color on column L and M
            for cols in ws2.iter_cols(min_col=12, max_col=13, min_row=2, max_row=ws2_max_row):
                for cell in cols:
                    cell.fill = PatternFill(start_color="FFF2CC", end_color="FFF2CC", fill_type="solid")
    
            # lock property data on ws2
            ws2.protection.sheet = True
    
            for col in ['L', 'M']:
                for cell in ws2[col]:
                    cell.protection = Protection(locked=False)
    
            # save workbook to save all changes made
            wb.save(os.path.join(path, file))
    
            print(f'{file} complete')
    
        else:
            continue
